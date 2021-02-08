# ————————————————
# @Time    : 2020/11/23 19:35
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : class_练习practice.py
# ————————————————

from openpyxl import load_workbook


class Practice:
    # 1、使用字典推导式将下面字符串格式的数据，改成自动化类型的数据
    cook_str = 'BODFSDJO=D890313503495;DJFL=32143541134;EOWJO=39851934827;LDKOWEI=342229082345'
    dic_str = {i.split("=")[0]: i.split("=")[1] for i in cook_str.split(";")}

    # 2、当前有文件case.excel，设计程序将excel中的用例读取到一个生成器？
    def gen_red (self, file_path, sheet_name):
        """
        读取excel数据的方法
        :param file_path: 文件路径
        :param sheet_name: 表单名称
        :return: 读取的数据
        """
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        row_max = sheet.max_row  # 最大行数
        column_max = sheet.max_column  # 最大列数

        for row in range(2, row_max + 1):
            row_dict = {}  # 每行数据
            for col in range(1, column_max + 1):
                # key 等于表头 value具体值
                row_dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col)

    # 3、请描述什么是可迭代对象？ 什么是迭代器？ 迭代器和生成器的区别？
    #  （1）可迭代对象是序列，可被for循环便利取值
    #  （2）Iterator接口提供了很多对集合元素进行迭代的方法。每一个集合类都包括了可以返回迭代器实例的迭代方法。
    #   迭代器可以在迭代过程中删除底层集合的元素，但是不可以直接调用集合的remove(Object obj)删除，可以通过迭代器的remove()方法删除
    #   为了方便的处理集合中的元素, Java中出现了一个对象, 该对象提供了一些方法专门处理集合中的元素.例如删除和获取集合中的元素.该对象就叫做迭代器(Iterator).对Collection进行迭代的类，称其为迭代器。
    #  （3）
    ''' 迭代器是一个更抽象的概念，任何对象，如果它的类有 next 方法和 iter 方法返回自己本身，对于 string、list、  # dict、tuple 等这类容器对象，使用 for 循环遍历是很方便的。在后台 for 语句对容器对象调用 iter()函数，iter()  
    是 python 的内置函数。iter()会返回一个定义了 next()方法的迭代器对象，它在容器中逐个访问容器内元素，next()也是 python 的内置函数。在没有后续元素时，next()会抛出一个 StopIteration 异常。
    生成器（Generator）是创建迭代器的简单而强大的工具。它们写起来就像是正规的函数，只是在需要返回数  # 据的时候使用 yield 语句。每次 next()被调用时，生成器会返回它脱离的位置（它记忆语句最后一次执行的位置和所有的数据值）
    区别：生成器能做到迭代器能做的所有事,而且因为自动创建了 iter()和 next()方法,生成器显得特别简洁,而且生成器也是高效的，使用生成器表达式取代列表解析可以同时节省内存。除了创建和保存程序状态的自动方法,
    当发生器终结时,还会自动抛出 StopIteration 异常。'''


# 示例
gen_case = ([i, j] for i in range(1, 5) for j in range(1, 5))
print(list(gen_case))


# 第二题老师方法
def read_excel (file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    col = ws.max_column + 1
    row = ws.max_row + 1
    test_data = ({ws.cell(1, j).value: ws.cell(i, j).value for j in range(1, col)} for i in range(2, row))
    return test_data


res = read_excel('interface_cases.xlsx', 'Sheet2')
# print(list(res))
for i in res:
    print(i)

if __name__ == '__main__':
    # first
    print(Practice.dic_str)
